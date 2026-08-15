#!/usr/bin/env python3
"""
Builds js/data.js for the UN Funds Map from
'South Sudan UN Funds - Reconstructed Dataset v3.xlsx' -> 'Master Data' sheet.

Master Data is already filtered by the source team to "active / committed"
projects only (confirmed against the raw per-fund sheets in the same workbook:
RSRTF excludes 3 already-closed project phases (~$38M), PBF excludes a pipeline
item and an approved-but-undisbursed item (~$9.6M), QIPs excludes
Completed/Cancelled/Not Started rows). This script does not re-filter; it
takes Master Data rows as the source of truth and reshapes them for the map.

Judgment calls made here (all flagged with `provisional: true` and/or surfaced
in `meta[fund].unattributed` in the output so the UI can label them):

- PBF: "Average Project value" column is already the correct per-county $ share
  (project value / number of counties it spans). Used directly.
- CERF: each row lists many real counties in one prose string sharing one
  project value -> split evenly across the listed counties. One row (Common
  Services - Humanitarian Air Services, $13M) lists mostly payam/airstrip
  names that are not real ADM2 counties -> left unattributed (fund total
  still includes it; not plotted to any single county).
- RSRTF: County column holds project-phase labels, not real counties. Mapped
  to real counties using the 5-area description supplied by the user, split
  evenly within each area. The Kong Koc Phase II area (Greater Tonj and
  adjacent Warrap/Lakes) doesn't have explicit county names in that
  description; mapped to Tonj East/North/South, Cueibet, Rumbek North
  (same counties the prior dashboard version used for this RSRTF corridor).
  ALL RSRTF county rows are marked provisional pending RSRTF's confirmation.
- PBF "To be determined" / "Nationwide" rows, and QIPs "Likuangole" ->
  mapped to Pibor (nearest real ADM2 county; Likuangole is a payam of Pibor),
  and "Ruweng Administrative Area (RRA)" -> mapped to Abiemnhom (nearest real
  ADM2 county) are left unattributed / best-effort respectively — see NOTES
  dict below.
"""
import json
import re
import openpyxl
from collections import defaultdict, OrderedDict

SRC = r"C:\Users\123\Downloads\UN Funds\South Sudan UN Funds - Reconstructed Dataset v3.xlsx"
OUT = r"C:\Users\123\Downloads\UN Funds\UN Funds Map\js\data.js"

CANON_COUNTIES = [
    'Abiemnhom', 'Abyei', 'Akobo', 'Aweil Centre', 'Aweil East', 'Aweil North',
    'Aweil South', 'Aweil West', 'Awerial', 'Ayod', 'Baliet', 'Bor South', 'Budi',
    'Canal/Pigi', 'Cueibet', 'Duk', 'Ezo', 'Fangak', 'Fashoda', 'Gogrial East',
    'Gogrial West', 'Guit', 'Ibba', 'Ikotos', 'Juba', 'Jur River', 'Kajo-keji',
    'Kapoeta East', 'Kapoeta North', 'Kapoeta South', 'Koch', 'Lafon', 'Lainya',
    'Leer', 'Longochuk', 'Luakpiny/Nasir', 'Maban', 'Magwi', 'Maiwut', 'Malakal',
    'Manyo', 'Maridi', 'Mayendit', 'Mayom', 'Melut', 'Morobo', 'Mundri East',
    'Mundri West', 'Mvolo', 'Nagero', 'Nyirol', 'Nzara', 'Panyijiar', 'Panyikang',
    'Pariang', 'Pibor', 'Pochalla', 'Raja', 'Renk', 'Rubkona', 'Rumbek Centre',
    'Rumbek East', 'Rumbek North', 'Tambura', 'Terekeka', 'Tonj East', 'Tonj North',
    'Tonj South', 'Torit', 'Twic', 'Twic East', 'Ulang', 'Uror', 'Wau', 'Wulu',
    'Yambio', 'Yei', 'Yirol East', 'Yirol West',
]
CANON_SET = {c.lower() for c in CANON_COUNTIES}

# Known raw-name -> canonical-name fixes (typos, "County" suffixes, reordering)
COUNTY_ALIASES = {
    'malakal county': 'Malakal',
    'jur river county': 'Jur River',
    'gogrial west county': 'Gogrial West',
    'gogrial\xa0west county': 'Gogrial West',
    'gogrial east county': 'Gogrial East',
    'gogrial\xa0east county': 'Gogrial East',
    'wau municipality': 'Wau',
    'pigi/canal': 'Canal/Pigi',
    'rumbek center county': 'Rumbek Centre',
    'rumbek centre county': 'Rumbek Centre',
    'rumbek north county': 'Rumbek North',
    'rumbek east county': 'Rumbek East',
    'cueibet county': 'Cueibet',
    'yiro west county': 'Yirol West',  # typo for "Yirol West"
    'likuangole': 'Pibor',  # payam of Pibor county
    'ruweng administrative area (rra)': 'Abiemnhom',  # best-effort: Ruweng AA ~ Abiemnhom/Pariang
}

# Payam / airstrip / camp / town -> real ADM2 county, for the CERF "Common
# Services - Humanitarian Air Services" row, whose location list mixes real
# county names with sub-county places. Each mapping was checked against a
# public source (OCHA, ReliefWeb, UNHCR, Wikipedia, CSRF county profiles) --
# see the chat record for citations. Two entries (marked LOW) could not be
# confirmed from any source found and are a best-effort guess; everything
# else is MEDIUM-HIGH or better confidence.
CERF_AIRSERVICES_LOCATION_TO_COUNTY = {
    'boma': 'Pibor',                # town in Pibor County (GPAA)
    'bor': 'Bor South',              # Bor town, HQ of Bor South County
    'nasir': 'Luakpiny/Nasir',       # Nasir town, in Luakpiny/Nasir County
    'chuil': 'Nyirol',               # Chuil Payam, Nyirol County
    'duk padiet': 'Duk',             # payam of Duk County
    'duk poktap': 'Duk',             # payam of Duk County
    'kurwai': 'Canal/Pigi',          # Kurwai Payam, Canal/Pigi County
    'mabior': 'Twic East',           # Mabior airstrip, Twic East County -- MEDIUM confidence
    'motot': 'Uror',                 # payam of Uror County
    'new fangak': 'Fangak',          # town in Fangak County
    'paguer lz': 'Fangak',           # Paguer landing zone/airstrip, Fangak County
    'pieri': 'Uror',                 # payam of Uror County
    'toch': 'Fangak',                # LOW confidence -- grouped with Fangak/Ayod/Paguer in UNHAS flight
                                      # resumption notices, but no source confirms its county outright
    'walgak': 'Akobo',               # town/payam in Akobo County
    'yuai': 'Uror',                  # payam of Uror County
    'agok': 'Abyei',                 # main town of the Abyei Administrative Area
    'kapoeta': 'Kapoeta South',      # Kapoeta town is the HQ of Kapoeta South County
    'nimule': 'Magwi',               # town in Magwi County
    'rumbek': 'Rumbek Centre',       # Rumbek town, HQ of Rumbek Centre County
    'aweil': 'Aweil Centre',         # Aweil town, HQ of Aweil Centre County
    'ajuong thok': 'Pariang',        # refugee camp in Pariang County
    'ganyiel': 'Panyijiar',          # payam/town, Panyijiar County
    'mankien': 'Mayom',              # payam/town, Mayom County
    'yida': 'Pariang',               # refugee camp in Pariang County
    'baramach': 'Ulang',             # "Barmach" Payam, Ulang County (alt. transliteration)
    'mathiang': 'Bor South',         # boma in Baidit Payam, Bor County
    'kuajok': 'Gogrial West',        # Kuajok town, HQ of Gogrial West County (Warrap State capital)
}

FUND_META = OrderedDict([
    ('SSHF', {'fullName': 'South Sudan Humanitarian Fund', 'color': '#2a78d6', 'type': 'Humanitarian pooled fund',
               'desc': 'Country-based pooled fund allocating to NGOs and UN agencies for time-critical, underfunded humanitarian priorities.'}),
    ('QIPs', {'fullName': 'Quick Impact Projects', 'color': '#D946EF', 'type': 'Mission quick-impact fund',
               'desc': 'UNMISS-funded small-scale, fast-delivery projects implemented by local and national NGOs to build confidence and support peace consolidation at the community level.'}),
    ('CERF', {'fullName': 'Central Emergency Response Fund', 'color': '#e0a800', 'type': 'Global emergency fund',
               'desc': 'UN global emergency fund providing rapid, up-front funding for new or rapidly deteriorating humanitarian crises.'}),
    ('PBF', {'fullName': 'Peacebuilding Fund', 'color': '#8b4513', 'type': 'Peace & stabilization fund',
              'desc': 'Supports peacebuilding, reconciliation and stabilization initiatives, often bridging humanitarian, development and peace efforts.'}),
    ('RSRTF', {'fullName': 'Reconciliation, Stabilization and Resilience Trust Fund', 'color': '#008300', 'type': 'Stabilization trust fund',
                'desc': 'Multi-partner trust fund supporting local reconciliation, stabilization and resilience-building in conflict-affected areas.'}),
    ('WPHF', {'fullName': "Women's Peace and Humanitarian Fund", 'color': '#e87ba4', 'type': 'Women, peace & humanitarian fund',
               'desc': "Funds women-led and women's rights organizations advancing the Women, Peace and Security and humanitarian agendas."}),
])
ORDER = ['CERF', 'PBF', 'QIPs', 'RSRTF', 'SSHF', 'WPHF']

# RSRTF provisional area -> county disaggregation (see module docstring)
RSRTF_AREAS = {
    'central equatoria phase ii': ['Kajo-keji', 'Yei', 'Morobo', 'Lainya'],
    'southern unity': ['Leer', 'Mayendit', 'Panyijiar'],
    'kong koc phase ii': ['Tonj East', 'Tonj North', 'Tonj South', 'Cueibet', 'Rumbek North'],
}

def clean_county(raw):
    if not raw:
        return None
    s = str(raw).replace('\xa0', ' ').strip()
    s = re.sub(r'\s+', ' ', s)
    key = s.lower().rstrip('.')
    key = re.sub(r'\s*/\s*', '/', key)  # "Luakpiny / Nasir" -> "luakpiny/nasir"
    if key in COUNTY_ALIASES:
        return COUNTY_ALIASES[key]
    if key in CANON_SET:
        # return canonical-cased version
        for c in CANON_COUNTIES:
            if c.lower() == key:
                return c
    return s  # unrecognized; caller decides what to do

def clean_partners(raw):
    if not raw:
        return []
    parts = re.split(r',\s*', str(raw).replace('\xa0', ' ').strip())
    out = []
    for p in parts:
        p = p.strip().rstrip('.')
        if p and p.upper() not in ('AND',):
            out.append(p)
    return out

def split_respecting_parens(s):
    """Splits on commas, but not commas inside (...) -- e.g.
    "Protection (CP, GBV), Health" -> ["Protection (CP, GBV)", "Health"]."""
    parts = []
    depth = 0
    buf = []
    for ch in s:
        if ch == '(':
            depth += 1
            buf.append(ch)
        elif ch == ')':
            depth = max(0, depth - 1)
            buf.append(ch)
        elif ch == ',' and depth == 0:
            parts.append(''.join(buf))
            buf = []
        else:
            buf.append(ch)
    if buf:
        parts.append(''.join(buf))
    return parts

def clean_sectors(raw):
    if not raw:
        return []
    parts = split_respecting_parens(str(raw).replace('\xa0', ' ').strip())
    out = []
    for p in parts:
        p = p.strip()
        if p:
            out.append(p)
    return out

def split_cerf_counties(raw, extra_aliases=None):
    """Splits a CERF prose county list into real ADM2 county names.
    `extra_aliases` (lowercase place -> county) resolves sub-county payams,
    airstrips, camps and towns that aren't themselves ADM2 counties. Returns
    None if a location still can't be resolved -- e.g. an unrecognized name."""
    s = str(raw).replace('\xa0', ' ').strip()
    s = re.sub(r'\s+and\s+', ', ', s, flags=re.IGNORECASE)
    s = s.rstrip('.')
    raw_names = [p.strip() for p in s.split(',') if p.strip()]
    cleaned = []
    for n in raw_names:
        n = re.sub(r'\s*\([^)]*\)', '', n).strip()  # strip "(Minkaman)" etc.
        c = clean_county(n)
        if (c is None or c.lower() not in CANON_SET) and extra_aliases:
            c = extra_aliases.get(n.lower())
        if c is None or c.lower() not in CANON_SET:
            return None  # bail: still unresolved
        cleaned.append(c)
    # de-duplicate while preserving first-seen order (several raw locations
    # can map to the same county, e.g. Chuil/Duk Padiet/Duk Poktap -> Duk/Nyirol)
    seen = set()
    out = []
    for c in cleaned:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out or None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['Master Data']
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

    # fund -> county(canonical) -> {value, projects, partners:set, sectors:set, provisional:bool, projectIds:set}
    #
    # `projects` on a county row means "how many funded projects touch this
    # county" -- correct to display per county, but NOT safe to sum across
    # counties for a fund-wide total whenever one project spans several
    # counties (true for PBF/CERF/RSRTF): the same project would be counted
    # once per county it touches. To let the dashboard compute a correct
    # deduplicated total for any set of counties (unfiltered or filtered),
    # every project gets a unique ID that's attached to every county row it
    # touches; a fund's true project count is the size of the UNION of those
    # IDs across whichever counties are in view, not the sum of `projects`.
    county_data = defaultdict(lambda: defaultdict(lambda: {'value': 0.0, 'projects': 0, 'partners': set(), 'sectors': set(), 'provisional': False, 'projectIds': set()}))
    unattributed = defaultdict(lambda: {'value': 0.0, 'items': []})
    fund_totals = defaultdict(lambda: {'value': 0.0, 'projects': 0.0})

    # --- pass 1: SSHF, WPHF, QIPs (one row = one project entry at one county;
    # `n` bundles multiple projects at that single county, so no cross-county
    # ID sharing is possible -- still tagged with synthetic IDs for a uniform
    # dedup path in app.js) and PBF (project split across counties, tracked
    # via a group ID per project so every county in the split shares the
    # SAME id and doesn't get double-counted) ---
    pbf_group_counter = 0
    pbf_group_ids = []
    for i, r in enumerate(rows):
        fund, theme, st, county_raw, val, avg, n, sd, ed, partners_raw = r
        if fund not in ('SSHF', 'WPHF', 'QIPs', 'PBF'):
            continue
        val = val or 0
        avg = avg or 0
        n = n or 0
        partners = clean_partners(partners_raw)
        sectors = clean_sectors(theme)

        if fund == 'PBF':
            per_county_value = avg if avg else val
            n_row = 1 if n else 0
            if val:  # a populated "Project/Programme Value" marks the first row of a new project group
                pbf_group_counter += 1
                group_n = int(n) if n else 1
                pbf_group_ids = ['PBF-g%d-p%d' % (pbf_group_counter, k) for k in range(1, group_n + 1)]
            row_ids = pbf_group_ids
        else:
            per_county_value = val
            n_row = n
            row_ids = ['%s-r%d-p%d' % (fund, i, k) for k in range(1, int(n_row) + 1)] if n_row else []

        county_clean = clean_county(county_raw)

        if fund == 'PBF' and county_raw in ('To be determined', 'Nationwide', None):
            label = county_raw or st or 'Unspecified'
            unattributed['PBF']['value'] += per_county_value
            unattributed['PBF']['items'].append({'label': label, 'value': per_county_value, 'projects': n_row})
            fund_totals['PBF']['value'] += per_county_value
            continue

        if county_clean is None or county_clean.lower() not in CANON_SET:
            # Unrecognized / unmappable county string -> park as unattributed rather than guess
            if county_raw:
                unattributed[fund]['value'] += per_county_value
                unattributed[fund]['items'].append({'label': str(county_raw), 'value': per_county_value, 'projects': n_row})
            fund_totals[fund]['value'] += per_county_value
            continue

        entry = county_data[fund][county_clean]
        entry['value'] += per_county_value
        # PBF: each county in a split project counts as 1 "project touches this county"
        # marker (not divided) -- matches how the source sheet marks n=1 on every row.
        entry['projects'] += (1 if fund == 'PBF' else n_row)
        entry['partners'] |= set(partners)
        entry['sectors'] |= set(sectors)
        entry['projectIds'] |= set(row_ids)
        fund_totals[fund]['value'] += per_county_value
        if fund != 'PBF':
            fund_totals[fund]['projects'] += n_row
        # PBF project count is recomputed separately below from the raw
        # "Number of Projects" column (1 per project, not per split county).

    # PBF project count: sum the "Number of Projects" column directly (1 per project, blanks=0)
    pbf_project_count = 0
    for r in rows:
        if r[0] == 'PBF' and r[6]:
            pbf_project_count += r[6]
    fund_totals['PBF']['projects'] = pbf_project_count

    # --- pass 2: CERF (prose county lists, split evenly). Each row's `n`
    # bundles n distinct projects (e.g. n partner-executed grants), each of
    # which touches every county the row lists -- so n synthetic IDs per row,
    # all shared across that row's counties. ---
    for i, r in enumerate(rows):
        fund, theme, st, county_raw, val, avg, n, sd, ed, partners_raw = r
        if fund != 'CERF':
            continue
        val = val or 0
        n = n or 0
        partners = clean_partners(partners_raw)
        sectors = clean_sectors(theme)
        fund_totals['CERF']['value'] += val
        fund_totals['CERF']['projects'] += n
        row_ids = set('CERF-r%d-p%d' % (i, k) for k in range(1, int(n) + 1)) if n else set()

        counties = split_cerf_counties(county_raw, extra_aliases=CERF_AIRSERVICES_LOCATION_TO_COUNTY)
        if not counties:
            unattributed['CERF']['value'] += val
            unattributed['CERF']['items'].append({'label': str(theme)[:60], 'value': val, 'projects': n})
            unresolved = [p.strip() for p in re.sub(r'\s+and\s+', ', ', str(county_raw), flags=re.IGNORECASE).rstrip('.').split(',') if p.strip()]
            unresolved = [p for p in unresolved if (clean_county(p) or '').lower() not in CANON_SET and p.lower() not in CERF_AIRSERVICES_LOCATION_TO_COUNTY]
            if unresolved:
                print('CERF row could not be geocoded -- unresolved location(s): %s' % ', '.join(unresolved))
            continue
        share = val / len(counties)
        for c in counties:
            entry = county_data['CERF'][c]
            entry['value'] += share
            entry['projects'] += n  # each listed county gets the full project-count marker for this row
            entry['partners'] |= set(partners)
            entry['sectors'] |= set(sectors)
            entry['projectIds'] |= row_ids
            entry['provisional'] = True  # even split across listed counties is an estimate

    # --- pass 3: RSRTF (phase-label rows, mapped via RSRTF_AREAS, split evenly).
    # Each label-group's `projects` bundles that many distinct funding
    # tranches, each touching every county the area maps to. ---
    rsrtf_groups = defaultdict(lambda: {'value': 0.0, 'projects': 0, 'partners': set(), 'sectors': set()})
    for r in rows:
        fund, theme, st, county_raw, val, avg, n, sd, ed, partners_raw = r
        if fund != 'RSRTF':
            continue
        label = str(county_raw).replace('\xa0', ' ').strip().lower()
        g = rsrtf_groups[label]
        g['value'] += (val or 0)
        g['projects'] += (n or 0)
        g['partners'] |= set(clean_partners(partners_raw))
        g['sectors'] |= set(clean_sectors(theme))
        fund_totals['RSRTF']['value'] += (val or 0)
        fund_totals['RSRTF']['projects'] += (n or 0)

    for label, g in rsrtf_groups.items():
        counties = RSRTF_AREAS.get(label)
        slug = re.sub(r'[^a-z0-9]+', '', label)
        group_n = int(g['projects']) if g['projects'] else 1
        row_ids = set('RSRTF-%s-p%d' % (slug, k) for k in range(1, group_n + 1))
        if not counties:
            unattributed['RSRTF']['value'] += g['value']
            unattributed['RSRTF']['items'].append({'label': label, 'value': g['value'], 'projects': g['projects']})
            continue
        share = g['value'] / len(counties)
        for c in counties:
            entry = county_data['RSRTF'][c]
            entry['value'] += share
            entry['projects'] += g['projects']
            entry['partners'] |= g['partners']
            entry['sectors'] |= g['sectors']
            entry['projectIds'] |= row_ids
            entry['provisional'] = True

    # --- assemble output ---
    funds_out = OrderedDict()
    for fund in ORDER:
        county_rows = []
        for county, e in sorted(county_data[fund].items()):
            row = {
                'county': county,
                'projects': round(e['projects'], 2) if e['projects'] % 1 else int(e['projects']),
                'value': round(e['value'], 2),
                'partners': sorted(e['partners']),
                'sectors': sorted(e['sectors']),
                'projectIds': sorted(e['projectIds']),
            }
            if e['provisional']:
                row['provisional'] = True
            county_rows.append(row)
        funds_out[fund] = county_rows

    meta_out = OrderedDict()
    for fund in ORDER:
        m = dict(FUND_META[fund])
        m['totalValue'] = round(fund_totals[fund]['value'], 2)
        m['totalProjects'] = round(fund_totals[fund]['projects'], 2) if fund_totals[fund]['projects'] % 1 else int(fund_totals[fund]['projects'])
        if unattributed[fund]['value']:
            m['unattributed'] = {
                'value': round(unattributed[fund]['value'], 2),
                'projects': sum(it['projects'] for it in unattributed[fund]['items']),
                'items': [{'label': it['label'], 'value': round(it['value'], 2), 'projects': it['projects']} for it in unattributed[fund]['items']],
            }
        meta_out[fund] = m

    grand_total_value = sum(meta_out[f]['totalValue'] for f in ORDER)

    out = {
        'generatedFrom': "South Sudan UN Funds - Reconstructed Dataset v3.xlsx (Master Data)",
        'asOf': '2026-08-15',
        'grandTotalValue': round(grand_total_value, 2),
        'order': ORDER,
        'meta': meta_out,
        'funds': funds_out,
    }

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write("// Auto-generated by scripts/build_data.py from\n")
        f.write("// 'South Sudan UN Funds - Reconstructed Dataset v3.xlsx' (Master Data sheet).\n")
        f.write("// Fund + county + $ value + number-of-projects + partner + sector records for the UN Funds Map app.\n")
        f.write("// RSRTF county assignment and any row marked \"provisional\" is a best-effort estimate\n")
        f.write("// pending confirmation from the fund; see scripts/build_data.py docstring for method.\n")
        f.write("window.FUNDS_DATA = ")
        f.write(json.dumps(out, indent=2, ensure_ascii=False))
        f.write(";\n")

    # ---- console summary for review ----
    print("Grand total value: ${:,.0f}".format(grand_total_value))
    for fund in ORDER:
        m = meta_out[fund]
        unattr = m.get('unattributed', {}).get('value', 0)
        print(f"{fund}: ${m['totalValue']:,.0f}  ({m['totalProjects']} projects, {len(funds_out[fund])} counties, ${unattr:,.0f} unattributed)")

    # ---- validation: every county name must exist in the canonical list ----
    bad = []
    for fund in ORDER:
        for row in funds_out[fund]:
            if row['county'].lower() not in CANON_SET:
                bad.append((fund, row['county']))
    if bad:
        print("\n!! UNRECOGNIZED COUNTY NAMES (will not plot on map):")
        for fund, c in bad:
            print(" ", fund, repr(c))
    else:
        print("\nAll county names validated against the 79-county topojson list.")

    # ---- validation: union of projectIds across all counties must equal the
    # authoritative per-fund project count (i.e. the ID-dedup path the
    # dashboard uses gives the same answer as the direct column sum) ----
    print()
    id_mismatch = False
    for fund in ORDER:
        all_ids = set()
        naive_sum = 0
        for row in funds_out[fund]:
            all_ids |= set(row['projectIds'])
            naive_sum += row['projects']
        # unattributed items aren't tagged with IDs in any county row (by
        # construction) -- back their project count out for a fair comparison
        unattr_projects = meta_out[fund].get('unattributed', {}).get('projects', 0)
        expected = meta_out[fund]['totalProjects'] - unattr_projects
        ok = len(all_ids) == expected
        if not ok:
            id_mismatch = True
        print(f"{fund}: naive county-sum={naive_sum}  ID-dedup total={len(all_ids)}  expected(fund total minus unattributed)={expected}  {'OK' if ok else 'MISMATCH'}")
    if id_mismatch:
        print("\n!! ID-dedup validation failed for at least one fund -- investigate before shipping.")
    else:
        print("\nID-dedup project counts validated against authoritative fund totals for every fund.")

if __name__ == '__main__':
    main()
